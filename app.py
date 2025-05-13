import os
import time
import uuid
import random
import string
import logging

from flask import (
    Flask, render_template,send_file, request,session,redirect,
    make_response, url_for
)
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_talisman import Talisman
from captcha.image import ImageCaptcha
from captcha.audio import AudioCaptcha
import numpy as np
import cv2
import openpyxl
import requests
import sentry_sdk
from sentry_sdk.integrations.flask import FlaskIntegration

from pydub import AudioSegment
from gtts import gTTS
from cryptography.hazmat.primitives.asymmetric import rsa
from cryptography.hazmat.primitives import serialization

# ──────────────────────────────────────────────────────────────────────────────
# App & Config
# ──────────────────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'ChangeThisSecretKey')

# Sentry for error & performance monitoring
sentry_sdk.init(
    dsn=os.environ.get('SENTRY_DSN'),
    integrations=[FlaskIntegration()],
    traces_sample_rate=0.1,
    environment=os.environ.get('FLASK_ENV', 'production')
)

# HTTPS redirect + strict CSP
csp = {
    'default-src': ["'self'"],
    'script-src': ["'self'", 'https://www.google.com/recaptcha/'],
    'style-src':  ["'self'"],
    'img-src':    ["'self'", 'data:'],
    'media-src':  ["'self'"]
}
Talisman(
    app,
    content_security_policy=csp,
    force_https=True,
    strict_transport_security=True,
    strict_transport_security_max_age=31536000
)

# Rate Limiting (Redis backend via REDIS_URL)
app.config['RATELIMIT_STORAGE_URL'] = os.environ.get('REDIS_URL', 'memory://')
limiter = Limiter(
    key_func=get_remote_address,
    default_limits=["10 per minute"]
)
limiter.init_app(app)

# Google reCAPTCHA v2 keys
RECAPTCHA_SITE_KEY   = os.environ.get('RECAPTCHA_SITE_KEY', '')
RECAPTCHA_SECRET_KEY = os.environ.get('RECAPTCHA_SECRET_KEY', '')

# Excel files for session storage & logs
SESSIONS_FILE = 'sessions.xlsx'
LOG_FILE      = 'captcha_logs.xlsx'
for fname, headers in [
    (SESSIONS_FILE, ['session_id','captcha','timestamp','used']),
    (LOG_FILE,      ['timestamp','ip','event','detail'])
]:
    if not os.path.exists(fname):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(fname)

# App-level logger
logging.basicConfig(
    filename='app.log',
    level=logging.INFO,
    format='%(asctime)s %(levelname)s: %(message)s'
)

quotes = [
  "Trust the process, not the outcome.",
  "Even bots need a break from logic.",
  "Security is humanity’s best puzzle.",
  "Only a human reads this far."
]
ai_quote = random.choice(quotes)

# ──────────────────────────────────────────────────────────────────────────────
# Excel helpers
# ──────────────────────────────────────────────────────────────────────────────
def store_session(sid, text):
    wb = openpyxl.load_workbook(SESSIONS_FILE)
    ws = wb.active
    ws.append([sid, text, time.time(), False])
    wb.save(SESSIONS_FILE)

def get_session(sid):
    wb = openpyxl.load_workbook(SESSIONS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == sid:
            return {'captcha': row[1], 'timestamp': row[2], 'used': row[3]}
    return None

def mark_used(sid):
    wb = openpyxl.load_workbook(SESSIONS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == sid:
            row[3].value = True
            break
    wb.save(SESSIONS_FILE)

def log_event(ip, event, detail=''):
    wb = openpyxl.load_workbook(LOG_FILE)
    ws = wb.active
    ws.append([time.time(), ip, event, detail])
    wb.save(LOG_FILE)
    logging.info(f"{ip} - {event}: {detail}")

def _init_session():
    sid = str(uuid.uuid4())
    resp = make_response()  # you’ll attach the template later
    resp.set_cookie('captcha_sid', sid, max_age=300)
    return sid

def _gen_text():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

def _fail(event, ip):
    log_event(ip, event)
    return "❌ Failed", 400

def _success():
    return "✅ Passed", 200

def generate_rsa_keypair():
    private_key = rsa.generate_private_key(public_exponent=65537, key_size=2048)
    private_bytes = private_key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.TraditionalOpenSSL,
        encryption_algorithm=serialization.NoEncryption()
    )
    public_key = private_key.public_key()
    public_bytes = public_key.public_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PublicFormat.SubjectPublicKeyInfo
    )
    return private_bytes.decode(), public_bytes.decode()
def hash_pw(pw): return hashlib.sha256(pw.encode()).hexdigest()

def load_db(): return openpyxl.load_workbook(DB)

def gen_rsa():  # returns (pub, priv)
    priv = rsa.generate_private_key(65537,2048)
    p_bytes = priv.private_bytes(
        serialization.Encoding.PEM,
        serialization.PrivateFormat.PKCS8,
        serialization.NoEncryption())
    pub = priv.public_key().public_bytes(
        serialization.Encoding.PEM,
        serialization.PublicFormat.SubjectPublicKeyInfo)
    return pub.decode(), p_bytes.decode()

def log_threat(ip, thr, det=''):
    wb = load_db(); ws=wb['threat_logs']
    ws.append([time.time(),ip,thr,det]); wb.save(DB)

def verify_recaptcha(token):
    resp = requests.post('https://www.google.com/recaptcha/api/siteverify',
                         data={'secret':RECAPTCHA_SECRET,'response':token}).json()
    return resp.get('success',False)
# ──────────────────────────────────────────────────────────────────────────────
# Bot-pattern detection (simple blur-variance check)
# ──────────────────────────────────────────────────────────────────────────────
def is_bot_image(image_bytes):
    arr = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_GRAYSCALE)
    lap = cv2.Laplacian(img, cv2.CV_64F).var()
    return lap < 50  # low variance = possible bot

# ──────────────────────────────────────────────────────────────────────────────
# Routes
# ──────────────────────────────────────────────────────────────────────────────
# 1) Registration & Key Generation
@app.route('/register', methods=['POST'])
def register():
    user = request.form['username']; pw = request.form['password']
    wb = load_db(); ws=wb['users']
    for r in ws.iter_rows(min_row=2,values_only=True):
        if r[0]==user: return "Username exists",400
    pub,priv = gen_rsa()
    ws.append([user,hash_pw(pw),pub,priv]); wb.save(DB)
    return render_template('keys.html', public_key=pub, private_key=priv)

# 2) Encrypt Message
@app.route('/encrypt', methods=['POST'])
def encrypt_msg():
    sender = request.form['username']
    msg    = request.form['message'].encode()
    wb=load_db(); ws=wb['users']
    keyrow = next(r for r in ws.iter_rows(min_row=2,values_only=True) if r[0]==sender)
    pub = serialization.load_pem_public_key(keyrow[2].encode())
    cipher = pub.encrypt(msg,
        serialization.OAEP(
            mgf=serialization.MGF1(algorithm=serialization.hashes.SHA256()),
            algorithm=serialization.hashes.SHA256(),label=None))
    b64 = cipher.hex()
    ws2=wb['messages']; ws2.append([sender,'',b64,time.time()]); wb.save(DB)
    return render_template('result.html', result=b64)

# 3) Decrypt Message
@app.route('/decrypt', methods=['POST'])
def decrypt_msg():
    user   = request.form['username']; cipher_hex=request.form['cipher']
    wb=load_db(); ws=wb['users']
    keyrow=next(r for r in ws.iter_rows(min_row=2,values_only=True) if r[0]==user)
    priv = serialization.load_pem_private_key(keyrow[3].encode(),None)
    plain = priv.decrypt(bytes.fromhex(cipher_hex),
        serialization.OAEP(
            mgf=serialization.MGF1(algorithm=serialization.hashes.SHA256()),
            algorithm=serialization.hashes.SHA256(),label=None))
    return render_template('result.html', result=plain.decode())

# 4) Threat Logger sample (logs every failed login)
@app.route('/login', methods=['POST'])
@limiter.limit("5/minute")
def login():
    user,pw = request.form['username'], request.form['password']
    wb=load_db(); ws=wb['users']
    row = next((r for r in ws.iter_rows(min_row=2,values_only=True) if r[0]==user),None)
    ip=get_remote_address()
    if not row or hash_pw(pw)!=row[1]:
        log_threat(ip,'login_failed',f"user={user}")
        return "Login failed",400
    session['user']=user
    return redirect(url_for('dashboard'))

# 5) Simple Dashboard
@app.route('/dashboard')
def dashboard():
    if 'user' not in session: return redirect(url_for('home'))
    wb=load_db()
    msgs = wb['messages']
    logs = wb['threat_logs']
    return render_template('dashboard.html',
        username=session['user'],
        messages=list(msgs.iter_rows(min_row=2,values_only=True)),
        threats=list(logs.iter_rows(min_row=2,values_only=True))
    )

@app.route('/')
@limiter.limit("500 per minute")
def index():
    # 1) Generate a new session ID and puzzle
    sid = str(uuid.uuid4())
    a, b = random.randint(1, 9), random.randint(1, 9)
    puzzle_q = f"What is {a} + {b}?"
    session['puzzle_ans'] = str(a + b)
    ai_quote = random.choice([
        "Trust the process, not the outcome.",
        "Even bots need a break from logic.",
        "Security is humanity’s best puzzle.",
        "Only a human reads this far."
    ])

    # 2) Store the CAPTCHA text
    text = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    store_session(sid, text)

    # 3) Generate and trim audio CAPTCHA
    os.makedirs('static', exist_ok=True)
    audio = AudioCaptcha()
    audio.write(text, f'static/{sid}.wav')
    # trim to 2 seconds (requires pydub & ffmpeg)
    sound = AudioSegment.from_wav(f'static/{sid}.wav')
    sound[:2000].export(f'static/{sid}.wav', format="wav")

    # 4) Build the response
    resp = make_response(render_template(
        'index.html',
        captcha_sid=sid,
        recaptcha_site_key=RECAPTCHA_SITE_KEY,
        puzzle_question=puzzle_q,
        ai_quote=ai_quote
    ))
    resp.set_cookie('captcha_sid', sid, max_age=300)
    return resp

@app.route('/generate_key')
def generate_key():
    priv, pub = generate_rsa_keypair()
    return render_template("keys.html", public_key=pub, private_key=priv)


@app.route('/captcha_image/<sid>')
def captcha_image(sid):
    sess = get_session(sid)
    if not sess or sess['used'] or time.time() - sess['timestamp'] > 300:
        return '', 404
    image = ImageCaptcha(width=280, height=90)
    return image.generate(sess['captcha']).read(), 200, {'Content-Type':'image/png'}

@app.route('/verify', methods=['POST'])
@limiter.limit("500 per minute")
def verify():
    ip = get_remote_address()

    # 1) Puzzle check
    user_puzzle = request.form.get('puzzle_answer', '').strip()
    if user_puzzle != session.get('puzzle_ans'):
        log_event(ip, 'puzzle_failed', f"got={user_puzzle}")
        return "❌ Puzzle incorrect", 400

    # 2) reCAPTCHA check
    recaptcha_resp = requests.post(
        'https://www.google.com/recaptcha/api/siteverify',
        data={
            'secret':   RECAPTCHA_SECRET_KEY,
            'response': request.form.get('g-recaptcha-response')
        }
    ).json()
    if not recaptcha_resp.get('success'):
        log_event(ip, 'recaptcha_failed', str(recaptcha_resp))
        return "reCAPTCHA failed", 400

    # 3) Session/CAPTCHA expiry check
    sid = request.cookies.get('captcha_sid')
    sess = get_session(sid)
    if not sess or sess['used'] or time.time() - sess['timestamp'] > 300:
        log_event(ip, 'captcha_expired', sid)
        return "Expired, reload", 400

    # 4) Bot-image detection
    file_storage = request.files.get('captcha_image')
    if file_storage and is_bot_image(file_storage.read()):
        log_event(ip, 'bot_detected', sid)
        return "Bot-like activity detected", 400

    # 5) CAPTCHA text check
    user_input = request.form.get('captcha_input', '').upper()
    if user_input == sess['captcha']:
        mark_used(sid)
        log_event(ip, 'success', sid)
        return "✅ Verified!"
    else:
        log_event(ip, 'captcha_failed', f"input={user_input}")
        return "❌ Incorrect", 400

@app.route("/generate-audio")
def generate_audio():
    text = "1234"
    sid = str(uuid.uuid4())
    audio = AudioCaptcha()

    # Make sure the static/ folder exists
    os.makedirs("static", exist_ok=True)

    # Save audio captcha
    audio.write(text, f'static/{sid}.wav')
    return f"Audio CAPTCHA saved as static/{sid}.wav"

@app.route('/select')
def select():
    """Let user choose CAPTCHA type."""
    return render_template('select.html')

@app.route('/captcha/audio')
def audio_captcha():
    sid      = _init_session()
    lang     = request.args.get('lang', 'en')       # 'zulu', 'xhosa', etc.
    text     = _gen_text()
    store_session(sid, text)

    # generate TTS audio
    tts = gTTS(text=text, lang=lang)
    os.makedirs(f'static/audio/{lang}', exist_ok=True)
    tts_path = f'static/audio/{lang}/{sid}.mp3'
    tts.save(tts_path)

    return render_template('audio.html',
        captcha_sid=sid,
        audio_url=url_for('static', filename=f'audio/{lang}/{sid}.mp3'),
        recaptcha_site_key=RECAPTCHA_SITE_KEY
    )

@app.route('/captcha/visual')
def visual_captcha():
    sid       = _init_session()
    category  = random.choice(['animals','colors','traffic_lights'])
    image_dir = f'static/captcha_images/{category}'
    images    = random.sample(os.listdir(image_dir), k=6)
    # mark which images are “correct” (e.g. animal) in session
    session['visual_answers'] = [img for img in images if category in img]

    return render_template('visual.html',
        captcha_sid=sid,
        images=[url_for('static', filename=f'captcha_images/{category}/{i}') for i in images],
        recaptcha_site_key=RECAPTCHA_SITE_KEY
    )

@app.route('/captcha/puzzle')
def puzzle_captcha():
    sid = _init_session()
    typ = random.choice(['math','dragdrop','sequence'])

    if typ == 'math':
        a,b = random.randint(1,9), random.randint(1,9)
        session['puzzle_ans'] = str(a+b)
        question = f"What is {a} + {b}?"
    elif typ == 'dragdrop':
        # assume you’ve pre-made draggable slices images under static/puzzles/
        parts = ['p1.png','p2.png','p3.png','p4.png']
        random.shuffle(parts)
        session['puzzle_ans'] = 'correct'  # validate client-side ordering
        question = parts
    else:  # sequence
        steps = ['step1.png','step2.png','step3.png']
        random.shuffle(steps)
        session['puzzle_ans'] = '1,2,3'
        question = steps

    return render_template('puzzle.html',
        captcha_sid=sid,
        type=typ,
        question=question,
        recaptcha_site_key=RECAPTCHA_SITE_KEY
    )

@app.route('/captcha/hybrid')
def hybrid_captcha():
    sid = _init_session()
    # reuse above logic
    # e.g. pick a math puzzle and audio in one page
    # …
    return render_template('hybrid.html')

@app.route('/validate', methods=['POST'])
def validate():
    sid = request.cookies.get('captcha_sid')
    ip  = get_remote_address()

    # 1) Puzzle (if present)
    if 'puzzle_answer' in request.form:
        if request.form['puzzle_answer'] != session.get('puzzle_ans'):
            return _fail('puzzle_failed', ip)

    # 2) Visual (if present)
    selected = request.form.getlist('visual_sel')
    if selected and set(selected) != set(session.get('visual_answers',[])):
        return _fail('visual_failed', ip)

    # 3) Audio/Image CAPTCHA text
    if 'captcha_input' in request.form:
        if request.form['captcha_input'].upper() != get_session(sid)['captcha']:
            return _fail('captcha_failed', ip)

    # 4) reCAPTCHA
    # … your existing reCAPTCHA validation …

    return _success()

if __name__=='__main__': app.run(host='0.0.0.0',port=5000)

